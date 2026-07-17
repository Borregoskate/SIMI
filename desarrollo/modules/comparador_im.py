"""
==============================================================
SIMI
Sistema Inteligente de Mercado e Investigaciones

comparador_im.py

Interfaz Streamlit del Comparador Inteligente por Procedimiento.

Responsabilidades:
- Permitir seleccionar un procedimiento existente.
- Ejecutar el análisis del procedimiento contra:
  * otros procedimientos operativos;
  * adjudicaciones históricas.
- Mostrar evolución interna por etapa.
- Mostrar referencia de mercado, tendencia, riesgo y recomendaciones.
- Generar un reporte temporal en Excel.

Este módulo:
- No carga archivos extraordinarios.
- No ejecuta SQL.
- No persiste resultados.
- Mantiene el resultado únicamente en session_state.

Autor: Jorge Saavedra
Versión: 2.0.0
==============================================================
"""

from decimal import Decimal
from io import BytesIO

import pandas as pd
import plotly.express as px
import streamlit as st

from services.comparador_im_service import (
    ComparadorIMService,
)


SESSION_RESULTADO = "comparador_im_resultado"
SESSION_PROCEDIMIENTO = "comparador_im_procedimiento"


# ==========================================================
# UTILIDADES
# ==========================================================

def _valor_entero(valor):
    try:
        return int(valor or 0)
    except (TypeError, ValueError, OverflowError):
        return 0


def _formatear_decimal(valor, decimales=2):
    if valor is None:
        return "Sin información"

    try:
        numero = Decimal(str(valor))
    except (TypeError, ValueError):
        return "Sin información"

    return f"{numero:,.{decimales}f}"


def _formatear_moneda(valor):
    if valor is None:
        return "Sin información"

    return f"${_formatear_decimal(valor)}"


def _formatear_porcentaje(valor):
    if valor is None:
        return "Sin información"

    return f"{_formatear_decimal(valor)}%"


def _dataframe(registros):
    if not registros:
        return pd.DataFrame()

    return pd.DataFrame(
        [
            dict(registro)
            for registro in registros
        ]
    )


def _convertir_decimales(df):
    if df.empty:
        return df

    resultado = df.copy()

    for columna in resultado.columns:
        resultado[columna] = resultado[columna].map(
            lambda valor: (
                float(valor)
                if isinstance(valor, Decimal)
                else valor
            )
        )

    return resultado


def _mostrar_estadisticas_mercado(
    titulo,
    estadisticas,
):
    """
    Muestra estadísticas de mercado con formato monetario.

    Si no existen observaciones, muestra un mensaje informativo
    en lugar de valores NULL.
    """
    st.write(f"### {titulo}")

    estadisticas = estadisticas or {}
    total_observaciones = _valor_entero(
        estadisticas.get("total_observaciones")
    )

    if total_observaciones == 0:
        st.info(
            "Sin datos suficientes para realizar un análisis "
            "estadístico."
        )
        return

    filas = [
        {
            "Indicador": "Observaciones",
            "Valor": total_observaciones,
        },
        {
            "Indicador": "Precio mínimo",
            "Valor": _formatear_moneda(
                estadisticas.get("precio_minimo")
            ),
        },
        {
            "Indicador": "Precio máximo",
            "Valor": _formatear_moneda(
                estadisticas.get("precio_maximo")
            ),
        },
        {
            "Indicador": "Precio promedio",
            "Valor": _formatear_moneda(
                estadisticas.get("precio_promedio")
            ),
        },
        {
            "Indicador": "Precio mediana",
            "Valor": _formatear_moneda(
                estadisticas.get("precio_mediana")
            ),
        },
        {
            "Indicador": "Desviación estándar",
            "Valor": _formatear_moneda(
                estadisticas.get("desviacion_estandar")
            ),
        },
        {
            "Indicador": "Percentil 25",
            "Valor": _formatear_moneda(
                estadisticas.get("percentil_25")
            ),
        },
        {
            "Indicador": "Percentil 75",
            "Valor": _formatear_moneda(
                estadisticas.get("percentil_75")
            ),
        },
        {
            "Indicador": "Rango intercuartil",
            "Valor": _formatear_moneda(
                estadisticas.get("rango_intercuartil")
            ),
        },
    ]

    st.dataframe(
        pd.DataFrame(filas),
        hide_index=True,
        width="stretch",
    )


# ==========================================================
# ESTADO
# ==========================================================

def _inicializar_estado():
    st.session_state.setdefault(SESSION_RESULTADO, None)
    st.session_state.setdefault(SESSION_PROCEDIMIENTO, None)


def _limpiar_resultado():
    st.session_state[SESSION_RESULTADO] = None
    st.session_state[SESSION_PROCEDIMIENTO] = None


# ==========================================================
# ENCABEZADO
# ==========================================================

def _mostrar_encabezado():
    st.header("Comparador Inteligente de Procedimientos")
    st.caption(
        "Analiza un procedimiento existente contra otros "
        "procedimientos operativos y el histórico acumulado "
        "de las mismas claves."
    )
    st.info(
        "El procedimiento seleccionado nunca se incluye dentro "
        "de su propia referencia de mercado."
    )


# ==========================================================
# SELECCIÓN Y EJECUCIÓN
# ==========================================================

def _cargar_procedimientos(service):
    try:
        return service.obtener_procedimientos_disponibles()
    except Exception as error:
        st.error(
            "No fue posible recuperar los procedimientos disponibles."
        )
        st.exception(error)
        return []


def _etiqueta_procedimiento(registro):
    numero = registro.get("numero_procedimiento")
    ejercicio = registro.get("ejercicio")
    total_claves = _valor_entero(
        registro.get("total_claves")
    )

    return (
        f"{numero} | Ejercicio: "
        f"{ejercicio if ejercicio is not None else 'S/A'} "
        f"| Claves: {total_claves}"
    )


def _mostrar_selector(service):
    procedimientos = _cargar_procedimientos(service)

    if not procedimientos:
        st.warning(
            "No existen procedimientos con claves registradas."
        )
        return None

    mapa = {
        _etiqueta_procedimiento(registro): registro
        for registro in procedimientos
    }

    seleccionado = st.selectbox(
        "Selecciona un procedimiento",
        options=list(mapa.keys()),
        key="comparador_im_selector_procedimiento",
    )

    procedimiento = mapa[seleccionado]

    descripcion = procedimiento.get("descripcion")

    if descripcion:
        st.caption(descripcion)

    return procedimiento


def _mostrar_acciones(procedimiento):
    columna_1, columna_2 = st.columns([3, 1])

    ejecutar = columna_1.button(
        "Ejecutar análisis inteligente",
        type="primary",
        width="stretch",
        disabled=procedimiento is None,
    )

    limpiar = columna_2.button(
        "Limpiar",
        width="stretch",
    )

    if limpiar:
        _limpiar_resultado()
        st.rerun()

    return ejecutar


def _ejecutar_analisis(service, procedimiento):
    try:
        with st.spinner(
            "Analizando el procedimiento contra el mercado "
            "operativo e histórico..."
        ):
            resultado = service.comparar_procedimiento(
                id_procedimiento=procedimiento[
                    "id_procedimiento"
                ]
            )
    except ValueError as error:
        st.warning(str(error))
        return None
    except Exception as error:
        st.error(
            "No fue posible completar el análisis del procedimiento."
        )
        st.exception(error)
        return None

    st.session_state[SESSION_RESULTADO] = resultado
    st.session_state[SESSION_PROCEDIMIENTO] = procedimiento[
        "id_procedimiento"
    ]

    return resultado


# ==========================================================
# INDICADORES
# ==========================================================

def _mostrar_indicadores(resultado):
    resumen = resultado.get("resumen", {}) or {}
    procedimiento = resultado.get("procedimiento", {}) or {}

    st.subheader(
        procedimiento.get("numero_procedimiento")
        or "Procedimiento"
    )

    st.caption(
        procedimiento.get("descripcion")
        or "Sin descripción"
    )

    fila_1 = st.columns(5)
    fila_1[0].metric(
        "Total de claves",
        _valor_entero(resumen.get("total_claves")),
    )
    fila_1[1].metric(
        "Con precio actual",
        _valor_entero(
            resumen.get("claves_con_precio_actual")
        ),
    )
    fila_1[2].metric(
        "Con referencia",
        _valor_entero(
            resumen.get("claves_con_referencia")
        ),
    )
    fila_1[3].metric(
        "Con histórico",
        _valor_entero(
            resumen.get("claves_con_historico")
        ),
    )
    fila_1[4].metric(
        "Riesgo alto",
        _valor_entero(
            resumen.get("claves_riesgo_alto")
        ),
    )

    fila_2 = st.columns(4)
    fila_2[0].metric(
        "Competitivas",
        _valor_entero(
            resumen.get("claves_competitivas")
        ),
    )
    fila_2[1].metric(
        "En mercado",
        _valor_entero(
            resumen.get("claves_en_mercado")
        ),
    )
    fila_2[2].metric(
        "Elevadas",
        _valor_entero(
            resumen.get("claves_elevadas")
        ),
    )
    fila_2[3].metric(
        "Sin referencia",
        _valor_entero(
            resumen.get("claves_sin_referencia")
        ),
    )


# ==========================================================
# TABLAS
# ==========================================================

def _mostrar_resumen_claves(registros):
    df = _dataframe(registros)

    if df.empty:
        st.info("No existen claves analizadas.")
        return

    columnas = {
        "clave": "Clave",
        "descripcion": "Descripción",
        "categoria": "Categoría",
        "etapa_actual": "Etapa actual",
        "precio_actual": "Precio actual",
        "precio_referencia": "Referencia mercado",
        "total_observaciones": "Observaciones",
        "variacion_porcentual": "Variación (%)",
        "clasificacion": "Clasificación",
        "tendencia": "Tendencia",
        "confianza": "Confianza",
        "riesgo": "Riesgo",
    }

    disponibles = [
        columna
        for columna in columnas
        if columna in df.columns
    ]

    st.dataframe(
        _convertir_decimales(
            df[disponibles].rename(columns=columnas)
        ),
        width="stretch",
        hide_index=True,
    )


def _mostrar_etapas(registros):
    df = _dataframe(registros)

    if df.empty:
        st.info("No existen etapas para mostrar.")
        return

    columnas = {
        "clave": "Clave",
        "mejor_precio_inicial": "Mejor inicial",
        "mejor_precio_viable": "Mejor viable",
        "mejor_precio_subasta": "Mejor subasta",
        "precio_adjudicado": "Precio adjudicado",
        "total_propuestas_iniciales": "Iniciales",
        "total_propuestas_viables": "Viables",
        "total_subastas": "Subastas",
        "total_adjudicaciones": "Adjudicaciones",
    }

    disponibles = [
        columna
        for columna in columnas
        if columna in df.columns
    ]

    st.dataframe(
        _convertir_decimales(
            df[disponibles].rename(columns=columnas)
        ),
        width="stretch",
        hide_index=True,
    )


def _mostrar_recomendaciones(registros):
    df = _dataframe(registros)

    if df.empty:
        st.info("No existen recomendaciones.")
        return

    columnas = {
        "clave": "Clave",
        "nivel": "Nivel",
        "titulo": "Recomendación",
        "mensaje": "Detalle",
    }

    disponibles = [
        columna
        for columna in columnas
        if columna in df.columns
    ]

    st.dataframe(
        df[disponibles].rename(columns=columnas),
        width="stretch",
        hide_index=True,
    )


# ==========================================================
# GRÁFICAS
# ==========================================================

def _mostrar_grafica_clasificacion(registros):
    df = _dataframe(registros)

    if df.empty or "clasificacion" not in df.columns:
        st.info("No existe información de clasificación.")
        return

    resumen = (
        df.groupby("clasificacion", dropna=False)
        .size()
        .reset_index(name="total")
    )

    figura = px.bar(
        resumen,
        x="clasificacion",
        y="total",
        title="Clasificación de claves",
        labels={
            "clasificacion": "Clasificación",
            "total": "Claves",
        },
    )

    st.plotly_chart(figura, width="stretch")


def _mostrar_grafica_riesgo(registros):
    df = _dataframe(registros)

    if df.empty or "riesgo" not in df.columns:
        st.info("No existe información de riesgo.")
        return

    resumen = (
        df.groupby("riesgo", dropna=False)
        .size()
        .reset_index(name="total")
    )

    figura = px.pie(
        resumen,
        names="riesgo",
        values="total",
        title="Riesgo por clave",
        hole=0.45,
    )

    st.plotly_chart(figura, width="stretch")


def _mostrar_grafica_etapas(registros):
    df = _dataframe(registros)

    if df.empty or "clave" not in df.columns:
        st.info("No existen etapas para graficar.")
        return

    columnas_precio = [
        "mejor_precio_inicial",
        "mejor_precio_viable",
        "mejor_precio_subasta",
        "precio_adjudicado",
    ]

    disponibles = [
        columna
        for columna in columnas_precio
        if columna in df.columns
    ]

    if not disponibles:
        st.info("No existen precios por etapa.")
        return

    claves = sorted(df["clave"].dropna().unique())

    clave = st.selectbox(
        "Clave para evolución interna",
        options=claves,
        key="comparador_im_clave_etapas",
    )

    fila = df[df["clave"] == clave].iloc[0]

    registros_etapas = []

    nombres = {
        "mejor_precio_inicial": "Inicial",
        "mejor_precio_viable": "Viable",
        "mejor_precio_subasta": "Subasta",
        "precio_adjudicado": "Adjudicación",
    }

    for columna in disponibles:
        valor = fila.get(columna)

        if pd.isna(valor):
            continue

        registros_etapas.append(
            {
                "etapa": nombres[columna],
                "precio": float(valor),
            }
        )

    if not registros_etapas:
        st.info(
            "La clave seleccionada no tiene precios por etapa."
        )
        return

    figura = px.line(
        pd.DataFrame(registros_etapas),
        x="etapa",
        y="precio",
        markers=True,
        title=f"Evolución interna — {clave}",
        labels={
            "etapa": "Etapa",
            "precio": "Precio unitario",
        },
    )

    st.plotly_chart(figura, width="stretch")


def _mostrar_grafica_distribucion(registros):
    df = _dataframe(registros)

    if df.empty:
        st.info(
            "No existe mercado operativo o histórico para graficar."
        )
        return

    requeridas = {"clave", "fuente", "precio"}

    if not requeridas.issubset(df.columns):
        st.info(
            "La distribución de mercado no está disponible."
        )
        return

    df["precio"] = pd.to_numeric(
        df["precio"],
        errors="coerce",
    )
    df = df.dropna(subset=["precio"])

    if df.empty:
        st.info("No existen precios numéricos.")
        return

    claves = sorted(df["clave"].dropna().unique())

    clave = st.selectbox(
        "Clave para distribución de mercado",
        options=claves,
        key="comparador_im_clave_distribucion",
    )

    filtrado = df[df["clave"] == clave]

    figura = px.box(
        filtrado,
        x="fuente",
        y="precio",
        points="all",
        title=f"Mercado operativo e histórico — {clave}",
        labels={
            "fuente": "Fuente",
            "precio": "Precio unitario",
        },
    )

    st.plotly_chart(figura, width="stretch")


def _mostrar_grafica_evolucion(registros):
    df = _dataframe(registros)

    if df.empty:
        st.info(
            "No existe una serie temporal suficiente."
        )
        return

    requeridas = {
        "clave",
        "precio_mediana",
        "numero_procedimiento",
    }

    if not requeridas.issubset(df.columns):
        st.info("La evolución histórica no está disponible.")
        return

    claves = sorted(df["clave"].dropna().unique())

    clave = st.selectbox(
        "Clave para evolución de mercado",
        options=claves,
        key="comparador_im_clave_evolucion",
    )

    filtrado = df[df["clave"] == clave].copy()
    filtrado["precio_mediana"] = pd.to_numeric(
        filtrado["precio_mediana"],
        errors="coerce",
    )
    filtrado = filtrado.dropna(
        subset=["precio_mediana"]
    )

    filtrado["periodo"] = filtrado.apply(
        lambda fila: (
            f"{fila.get('ejercicio') or 'S/A'} — "
            f"{fila.get('numero_procedimiento') or 'Sin procedimiento'}"
        ),
        axis=1,
    )

    figura = px.line(
        filtrado,
        x="periodo",
        y="precio_mediana",
        markers=True,
        title=f"Evolución operativa e histórica — {clave}",
        labels={
            "periodo": "Periodo",
            "precio_mediana": "Precio mediano",
        },
    )

    st.plotly_chart(figura, width="stretch")


# ==========================================================
# DETALLE POR CLAVE
# ==========================================================

def _mostrar_detalle_clave(claves):
    if not claves:
        st.info("No existen claves analizadas.")
        return

    mapa = {
        (
            f"{registro.get('clave')} — "
            f"{registro.get('descripcion') or 'Sin descripción'}"
        ): registro
        for registro in claves
    }

    etiqueta = st.selectbox(
        "Selecciona una clave",
        options=list(mapa.keys()),
        key="comparador_im_detalle_clave",
    )

    item = mapa[etiqueta]
    comparacion = item.get("comparacion", {}) or {}
    referencia = item.get("referencia", {}) or {}
    etapas = item.get("etapas_procedimiento", {}) or {}
    estadisticas = item.get(
        "estadisticas_mercado",
        {},
    ) or {}

    fila_1 = st.columns(5)
    fila_1[0].metric(
        "Etapa actual",
        item.get("etapa_actual") or "SIN PRECIO",
    )
    fila_1[1].metric(
        "Precio actual",
        _formatear_moneda(
            item.get("precio_actual")
        ),
    )
    fila_1[2].metric(
        "Referencia",
        _formatear_moneda(
            referencia.get("precio_referencia")
        ),
    )
    fila_1[3].metric(
        "Variación",
        _formatear_porcentaje(
            comparacion.get("variacion_porcentual")
        ),
    )
    fila_1[4].metric(
        "Riesgo",
        item.get("nivel_riesgo")
        or "INDETERMINADO",
    )

    st.write("### Evolución del procedimiento")

    df_etapas = pd.DataFrame(
        [
            {
                "Etapa": "Mejor precio inicial",
                "Precio": _formatear_moneda(
                    etapas.get("mejor_precio_inicial")
                ),
            },
            {
                "Etapa": "Mejor precio viable",
                "Precio": _formatear_moneda(
                    etapas.get("mejor_precio_viable")
                ),
            },
            {
                "Etapa": "Mejor precio subasta",
                "Precio": _formatear_moneda(
                    etapas.get("mejor_precio_subasta")
                ),
            },
            {
                "Etapa": "Precio adjudicado",
                "Precio": _formatear_moneda(
                    etapas.get("precio_adjudicado")
                ),
            },
        ]
    )

    st.dataframe(
        df_etapas,
        hide_index=True,
        width="stretch",
    )

    _mostrar_estadisticas_mercado(
        "Mercado operativo",
        estadisticas.get("mercado_operativo", {}),
    )

    _mostrar_estadisticas_mercado(
        "Mercado histórico",
        estadisticas.get("mercado_historico", {}),
    )

    _mostrar_estadisticas_mercado(
        "Mercado consolidado",
        estadisticas.get("mercado_consolidado", {}),
    )

    st.write("### Recomendaciones")

    recomendaciones = item.get("recomendaciones", [])

    if not recomendaciones:
        st.info("No existen recomendaciones.")
    else:
        for recomendacion in recomendaciones:
            nivel = recomendacion.get("nivel")
            titulo = recomendacion.get("titulo")
            mensaje = recomendacion.get("mensaje")

            st.markdown(
                f"**{nivel} — {titulo}**  \n{mensaje}"
            )


# ==========================================================
# DESCARGA TEMPORAL
# ==========================================================

def _generar_excel(resultado):
    salida = BytesIO()
    tablas = resultado.get("tablas", {}) or {}

    with pd.ExcelWriter(
        salida,
        engine="openpyxl",
    ) as writer:
        hojas = {
            "Resumen claves": tablas.get(
                "resumen_claves",
                [],
            ),
            "Etapas procedimiento": tablas.get(
                "etapas_procedimiento",
                [],
            ),
            "Recomendaciones": tablas.get(
                "recomendaciones",
                [],
            ),
            "Evolucion mercado": tablas.get(
                "evolucion_precios",
                [],
            ),
            "Distribucion mercado": tablas.get(
                "distribucion_mercado",
                [],
            ),
        }

        for nombre, registros in hojas.items():
            df = _convertir_decimales(
                _dataframe(registros)
            )

            if df.empty:
                df = pd.DataFrame(
                    [{"Información": "Sin registros"}]
                )

            # Eliminar columnas técnicas
            columnas_ocultar = [
                "id_procedimiento",
                "id_clave",
            ]

            df = df.drop(
                columns=[
                    columna
                    for columna in columnas_ocultar
                    if columna in df.columns
                ],
                errors="ignore",
            )

            df.to_excel(
                writer,
                sheet_name=nombre[:31],
                index=False,
            )

            ws = writer.sheets[nombre[:31]]

            from openpyxl.styles import numbers
            from openpyxl.utils import get_column_letter

            # Formato de moneda para columnas H e I
            for col in ("F", "G"):
                if ws.max_column >= ord(col) - 64:
                    for cell in ws[col][1:]:
                        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            # Formato porcentual para columna K
            if ws.max_column >= 11:
                for cell in ws["I"][1:]:
                    cell.number_format = "0.00%"

            # Ajustar automáticamente el ancho de las columnas
            for column_cells in ws.columns:
                longitud = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                ws.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = min(max(longitud + 2, 12), 45)

    salida.seek(0)
    return salida.getvalue()


def _mostrar_descarga(resultado):
    procedimiento = resultado.get(
        "procedimiento",
        {},
    ) or {}

    numero = (
        procedimiento.get("numero_procedimiento")
        or "procedimiento"
    )

    nombre = (
        f"{numero}_comparador_inteligente_simi.xlsx"
        .replace("/", "_")
        .replace("\\", "_")
    )

    st.download_button(
        "Descargar reporte temporal",
        data=_generar_excel(resultado),
        file_name=nombre,
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        width="stretch",
    )


# ==========================================================
# RESULTADO
# ==========================================================

def _mostrar_resultado(resultado):
    _mostrar_indicadores(resultado)

    tablas = resultado.get("tablas", {}) or {}

    st.divider()
    st.subheader("Visualización analítica")

    columna_1, columna_2 = st.columns(2)

    with columna_1:
        _mostrar_grafica_clasificacion(
            tablas.get("resumen_claves", [])
        )

    with columna_2:
        _mostrar_grafica_riesgo(
            tablas.get("resumen_claves", [])
        )

    columna_3, columna_4 = st.columns(2)

    with columna_3:
        _mostrar_grafica_etapas(
            tablas.get(
                "etapas_procedimiento",
                [],
            )
        )

    with columna_4:
        _mostrar_grafica_distribucion(
            tablas.get(
                "distribucion_mercado",
                [],
            )
        )

    st.divider()

    _mostrar_grafica_evolucion(
        tablas.get("evolucion_precios", [])
    )

    st.divider()
    st.subheader("Detalle del análisis")

    pestañas = st.tabs(
        [
            "Resumen por clave",
            "Etapas del procedimiento",
            "Recomendaciones",
            "Detalle de clave",
        ]
    )

    with pestañas[0]:
        _mostrar_resumen_claves(
            tablas.get("resumen_claves", [])
        )

    with pestañas[1]:
        _mostrar_etapas(
            tablas.get(
                "etapas_procedimiento",
                [],
            )
        )

    with pestañas[2]:
        _mostrar_recomendaciones(
            tablas.get("recomendaciones", [])
        )

    with pestañas[3]:
        _mostrar_detalle_clave(
            resultado.get("claves", [])
        )

    st.divider()
    _mostrar_descarga(resultado)


# ==========================================================
# CONTROLADOR
# ==========================================================

def mostrar_comparador_im():
    """
    Renderiza el Comparador Inteligente por Procedimiento.
    """
    _inicializar_estado()
    _mostrar_encabezado()

    service = ComparadorIMService()
    procedimiento = _mostrar_selector(service)

    ejecutar = _mostrar_acciones(procedimiento)

    if ejecutar and procedimiento is not None:
        resultado = _ejecutar_analisis(
            service=service,
            procedimiento=procedimiento,
        )
    else:
        resultado = st.session_state.get(
            SESSION_RESULTADO
        )

    if resultado:
        _mostrar_resultado(resultado)